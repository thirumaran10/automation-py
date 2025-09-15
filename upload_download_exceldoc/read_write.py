#Read & update the excel with new data

import openpyxl
book = openpyxl.load_workbook("T:\\vscode\\python_selenium\\data\\demo.xlsx")

sheet = book.active #to get control of sheet object
dict = {}
cell = sheet.cell(row=1, column=2)#assess row and column
print(cell.value)#extract the value from cell

#to write the name to excel
sheet.cell(row=2, column=2).value = "vinayak"
print(sheet.cell(row=2, column=2).value)

#total max no of rows and columns present in the sheet
print(sheet.max_row)
print(sheet.max_column)
print(sheet['a4'].value)#give the value of cell 

#to print all the datas in table
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

#print specific datas
for i in range(1,sheet.max_row+1):#to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":#the text matchs only it enter into the inner for loop and scan
        for j in range(1,sheet.max_column+1):#to get column
            print("specific data:", sheet.cell(row=i, column=j).value)

#insted of printing we store value to dictonary
dict = {}
for i in range(1,sheet.max_row+1):#to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":#the text matchs only it enter into the inner for loop and scan
        for j in range(2,sheet.max_column+1):#to get column
            #dict["lastname"]="mahadev"
            dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(dict)      


#update excel
import openpyxl

file_path = "T:\\vscode\\python_selenium\\data\\download.xlsx"

def update_excel_data(file_path, columnname, searchterm, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active#to get control of sheet object
    Dict = {}

    #to lock the price column
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == columnname: #price      
            Dict["col"] = i

    #searchterm of apple
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchterm: #Apple
                Dict["row"] = i

    #value upadate to 500
    sheet.cell(row= Dict["row"], column= Dict["col"]).value= new_value #500
    book.save(file_path)